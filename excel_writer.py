import threading
import datetime
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE
from openpyxl.styles import Font, Color, PatternFill
import json
import getpass

csv_labels = ['test_name', 'current_time', 'test_duration', 'test_started', 'test_finished', 'node_name', 'battery_soc_avg_mwh', 'battery_soc_max_mwh', 'battery_soc_min_mwh', 'battery_soc_unlimited_avg_mwh', 'battery_soc_unlimited_max_mwh', 'battery_soc_unlimited_min_mwh', 
 'battery_excess_input_sum', 'battery_renewable_input_sum', 'battery_energy_usage_sum', 'battery_status_failed_percent', 'battery_status_failed_longest_duration', 'battery_status_failed_total_duration', 'battery_status_failed_switch_to_state_counter',
'battery_status_pending_percent', 'battery_status_pending_longest_duration', 'battery_status_pending_total_duration', 'battery_status_pending_switch_to_state_counter', 'battery_status_running_percent', 'battery_status_running_longest_duration',
 'battery_status_running_total_duration', 'battery_status_running_switch_to_state_counter', 'power_usage_incremental', 'down_time_minute', 'down_time_percent', 'replacements', 'cpu_usage_avg', 'cpu_usage_max', 'cpu_usage_up_avg', 'cpu_usage_up_max', 'cpuFreq_avg',
'memory_usage_avg', 'memory_usage_max', 'bw_usage_sent_mb', 'bw_usage_recv_mb'     
 'app_splitter', 'app', 'created', 'sent_sum', 'sent_percent', 'code200_sum', 'code200_percent', 'code500', 'code502', 'code503', 'code-1', 'others', 'drop451_sum', 'drop451_percent', 'drop_boot452_sum', 'drop_boot452_percent', 
 'dropped_sensors_no_host453_sum', 'dropped_sensors_no_host453_percent', 'dropped_sensors_func_host_down454_sum','dropped_sensors_func_host_down454_percent','dropped_sensors_func_local_host_down455_sum','dropped_sensors_func_local_host_down455_percent', 
 'dropped_sensors_hiccup450_sum', 'dropped_sensors_hiccup450_percent', 'dropped_sensors_active_sensor449_sum', 'dropped_sensors_active_sensor449_percent',
 'adm_avg', 'adm_max', 'qu_avg', 'qu_max', 'exec_avg', 'exec_max', 'rt_suc_avg', 'rt_suc_max', 'rt_suc_fail_avg', 'rt_suc_fail_max', 'useless', 'throu2', 
 'p0_suc', 'p25_suc', 'p50_suc', 'p75_suc', 'p90_suc', 'p95_suc', 'p99_suc', 'p99.9_suc', 'p100_suc', 'p0_suc_fail', 'p25_suc_fail', 'p50_suc_fail', 'p75_suc_fail', 'p90_suc_fail', 'p95_suc_fail', 'p99_suc_fail', 'p99.9_suc_fail', 'p100_suc_fail', 'detect_sum', 'detect_avg', 'detect_accuracy',]   # csv_labels = ['test','time','dur','start','finish','name','battery_soc_avg','battery_soc_max','battery_soc_min','battery_soc_unlimited_avg','battery_soc_unlimited_max','battery_soc_unlimited_min',


def write_avg(rows_count, excel_file_path):
    
    try:
        #get workbook
        wb = load_workbook(filename = excel_file_path)
        #get worksheet
        sheet = wb['nodes']
    except Exception as e:
        return str(e)

    #last written row   index 
    max_row = sheet.max_row
    #current row index
    new_row = max_row + 1
    
    #name it
    sheet["A" + str(new_row)] = "avg"
    #for i in range (100):
    #    sheet["C" + str(new_row)].font = Font(bold=True)
    
    #get indexes
    first_row = str(max_row - (rows_count - 1))
    end_row = str(max_row)
    
    #use iter_cols() ?????
    for row in sheet.iter_cols(min_row=new_row, max_row=new_row, min_col=2, max_col=100):
        for cell in row:
            column_letter = cell.column_letter
            cell.value= "=AVERAGE(" + column_letter + first_row + ":" + column_letter + end_row +")"
            #bold it
            cell.font = Font(bold=True)
            #fill color
            cell.fill = PatternFill(fgColor="DCDCDC")
        
    #columnIds = ['C', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
    #calculate test duration avg 
    #for columnId in columnIds:
        #set value
        #sheet[columnId + str(new_row)] = "=AVERAGE(" + columnId + first_row + ":" + columnId + end_row +")"
        #bold it
        #sheet[columnId + str(new_row)].font = Font(bold=True)
        #fill color
        #sheet[columnId + str(new_row)].fill = PatternFill(fgColor="00FF00FF")
    
    #save
    wb.save(filename = excel_file_path)
    wb.close()

    return new_row

#scheduler sheet
def write_scheduler(metrics, excel_file_path):
    lock = threading.Lock()
    with lock:
        try:
            wb = load_workbook(filename = excel_file_path)
            sheet = wb['scheduler']
        except Exception as e:
            return str(e)
        #freeze first column and row
        sheet.freeze_panes = "B2"
        
        #write new row to the end
        max_row = sheet.max_row
        print('max_row= ' + str(max_row))
        
        #add scheduler info
        down_counter = metrics["scheduler"]["down_counter"]
        print(down_counter)
        for key, value in down_counter.items():
            row = []
            #test name A
            row.extend([metrics["info"]["test_name"]])            
            #timestamp B
            current_time = str(datetime.datetime.now(datetime.timezone.utc).astimezone())
            row.extend([current_time])
        
            #node name C
            row.extend([key])
            #down_counter D
            row.extend([value])
            
            #sum placements E
            row.extend([metrics["scheduler"]["placements"]["sum"]])
            #placements per worker F..
            row.extend([metrics["scheduler"]["placements"]["workers"][key]])
            
            #placements per functions
            #add placmement per function
            per_functions = metrics["scheduler"]["placements"]["functions"]
            for key, value in per_functions.items():
                #function name
                row.extend([key])
                #function placements
                row.extend([value])
                
            #append this node info
            sheet.append(row)  
            
            #stylish
            max_row = sheet.max_row
            current_row = max_row
            
            #bold node name
            sheet["C" + str(current_row)].font = Font(bold=True)
            
                            
        #save
        wb.save(filename = excel_file_path)
        wb.close()


#write metrics
def write(metrics, excel_file_path):
    result = "null"
    current_row = 0
    
    #write scheduler
    if "scheduler" in metrics:
        try:
            write_scheduler(metrics, excel_file_path)
        except Exception as e:
            print('write_scheduler:' + str(e))

    #write nodes  
    lock = threading.Lock()
    with lock:
        try:
            wb = load_workbook(filename = excel_file_path)
            sheet = wb['nodes']
        except Exception as e:
            return str(e)
        #freeze first column and row
        sheet.freeze_panes = "B2"
        
            
        #write new row to the end
        max_row = sheet.max_row
        print('max_row= ' + str(max_row))
        row = []

        #test name A
        row.extend([metrics["info"]["test_name"]])
        #timestamp B
        current_time = str(datetime.datetime.now(datetime.timezone.utc).astimezone())
        row.extend([current_time])

        #info  C D E 
        row.extend([metrics["info"]["test_duration"],
                  metrics["info"]["test_started"],
                  metrics["info"]["test_finished"]])
        print('info')
        print(row)

        #node F G H I J K L M N O P Q R S T U V W...
        row.extend([metrics["node"]["name"],
                  metrics["node"]["battery_soc_avg"],
                  metrics["node"]["battery_soc_max"],
                  metrics["node"]["battery_soc_min"],
                  metrics["node"]["battery_soc_unlimited_avg"],
                  metrics["node"]["battery_soc_unlimited_max"],
                  metrics["node"]["battery_soc_unlimited_min"],
                  metrics["node"]["battery_excess_input_sum"],
                  metrics["node"]["battery_renewable_input_sum"],
                  metrics["node"]["battery_energy_usage_sum"],
                  metrics["node"]["battery_status_failed_percent"],
                  metrics["node"]["battery_status_failed_longest_duration"],
                  metrics["node"]["battery_status_failed_total_duration"],
                  metrics["node"]["battery_status_failed_switch_to_state_counter"],
                  metrics["node"]["battery_status_pending_percent"],
                  metrics["node"]["battery_status_pending_longest_duration"],
                  metrics["node"]["battery_status_pending_total_duration"],
                  metrics["node"]["battery_status_pending_switch_to_state_counter"],
                  metrics["node"]["battery_status_running_percent"],
                  metrics["node"]["battery_status_running_longest_duration"],
                  metrics["node"]["battery_status_running_total_duration"],
                  metrics["node"]["battery_status_running_switch_to_state_counter"],
                  metrics["node"]["power_usage_incremental"],
                  metrics["node"]["down_time"]["minute"],
                  metrics["node"]["down_time"]["percent"],
                  metrics["node"]["replacements"],
                  metrics["node"]["cpu_usage"]["avg"],
                  metrics["node"]["cpu_usage"]["max"],
                  metrics["node"]["cpu_usage_up"]["avg"],
                  metrics["node"]["cpu_usage_up"]["max"],
                  metrics["node"]["cpuFreq"]["avg"],
                #   metrics["node"]["cpuFreq"]["min"],
                #   metrics["node"]["cpuFreq"]["max"],
                  metrics["node"]["memory_usage"]["avg"],
                  metrics["node"]["memory_usage"]["max"],
                  metrics["node"]["bw_usage"]["sent_mb"],
                  metrics["node"]["bw_usage"]["recv_mb"]])

        #splitter ?
        row.extend(['**end node**'])


        #apps ... Y Z AA ...
        #if node has been LOAD_GENERATOR OR STANDALONE
        if "app_order" in metrics:
            #and has had enabled app
            if len(metrics["app_order"]):
                for app in metrics["app_order"]:
                    print(metrics[app])
                    # --> app name: first app is "app_overall"
                    row.extend([app])

                    row.extend([metrics[app]["created"],
                              metrics[app]["sent"]["sum"],
                              metrics[app]["sent"]["percent"],
                              metrics[app]["code200"]["sum"],
                              metrics[app]["code200"]["percent"],
                              metrics[app]["code500"],
                              metrics[app]["code502"],
                              metrics[app]["code503"],
                              metrics[app]["code-1"],
                              metrics[app]["others"],
                              metrics[app]["dropped"]["sum"],
                              metrics[app]["dropped"]["percent"],
                              metrics[app]["dropped_in_bootup"]["sum"],
                              metrics[app]["dropped_in_bootup"]["percent"],
                              metrics[app]["dropped_sensors_no_host"]["sum"],
                              metrics[app]["dropped_sensors_no_host"]["percent"],
                              metrics[app]["dropped_sensors_func_host_down"]["sum"],
                              metrics[app]["dropped_sensors_func_host_down"]["percent"],
                              metrics[app]["dropped_sensors_func_local_host_down"]["sum"],
                              metrics[app]["dropped_sensors_func_local_host_down"]["percent"],
                              metrics[app]["dropped_sensors_hiccup"]["sum"],
                              metrics[app]["dropped_sensors_hiccup"]["percent"], 
                              metrics[app]["dropped_sensors_active_sensor"]["sum"],
                              metrics[app]["dropped_sensors_active_sensor"]["percent"], 
                              metrics[app]["admission_dur"]["avg"],
                              metrics[app]["admission_dur"]["max"],
                              metrics[app]["queue_dur"]["avg"],
                              metrics[app]["queue_dur"]["max"],
                              metrics[app]["exec_dur"]["avg"],
                              metrics[app]["exec_dur"]["max"],
                              metrics[app]["round_trip_suc"]["avg"],
                              metrics[app]["round_trip_suc"]["max"],
                              metrics[app]["round_trip_suc_fail"]["avg"],
                              metrics[app]["round_trip_suc_fail"]["max"],  
                              metrics[app]["useless_exec_dur"],
                              metrics[app]["throughput2"]["avg"],
                              metrics[app]["percentiles_suc"]["p0"],
                              metrics[app]["percentiles_suc"]["p25"],
                              metrics[app]["percentiles_suc"]["p50"],
                              metrics[app]["percentiles_suc"]["p75"],
                              metrics[app]["percentiles_suc"]["p90"],
                              metrics[app]["percentiles_suc"]["p95"],
                              metrics[app]["percentiles_suc"]["p99"],
                              metrics[app]["percentiles_suc"]["p99.9"],
                              metrics[app]["percentiles_suc"]["p100"],
                              metrics[app]["percentiles_suc_fail"]["p0"],
                              metrics[app]["percentiles_suc_fail"]["p25"],
                              metrics[app]["percentiles_suc_fail"]["p50"],
                              metrics[app]["percentiles_suc_fail"]["p75"],
                              metrics[app]["percentiles_suc_fail"]["p90"],
                              metrics[app]["percentiles_suc_fail"]["p95"],
                              metrics[app]["percentiles_suc_fail"]["p99"],
                              metrics[app]["percentiles_suc_fail"]["p99.9"],
                              metrics[app]["percentiles_suc_fail"]["p100"],
                              '-'.join(map(str, metrics[app]["executor_ips"]["hosts"]["ips"])),
                              '-'.join(map(str, metrics[app]["executor_ips"]["hosts"]["counter"])),
                              '-'.join(map(str, metrics[app]["executor_ips"]["pods"]["ips"])),
                              '-'.join(map(str, metrics[app]["executor_ips"]["pods"]["counter"])),
                              metrics[app]["detected_objects"]["sum"],
                              metrics[app]["detected_objects"]["avg"],
                              metrics[app]["detected_objects"]["accuracy_avg"]])

                    #splitter 
                    row.extend(['**end app**'])

        
        #append to excel
        sheet.append(row)
        # sheet.append(labels)
        
        #stylish the new row
        current_row = max_row + 1
        #bold test name
        sheet["A" + str(current_row)].font = Font(bold=True)
        #bold node_name F
        sheet["F" + str(current_row)].font = Font(bold=True)
        #bold if not master
        if not "master" in metrics["node"]["name"]:
            #down_time percent 
            sheet["I" + str(current_row)].font = Font(bold=True)
            #cpuUtil_up_avg
            sheet["L" + str(current_row)].font = Font(bold=True)
            
        #save
        wb.save(filename = excel_file_path)
        wb.close()

        #append to CSV file
        csv_file_path = excel_file_path.split('.')[0] + '.csv'
        csv_write(csv_file_path, [row])
        # csv_write(csv_file_path, labels)

        return 'success', current_row

#append to csv file
def csv_write(csv_file_path, data):
    import numpy as np
    #!/usr/bin/env python3

    csv_file = open(csv_file_path, 'ab')
    np.savetxt(csv_file, data, delimiter=",", fmt="%s")
    csv_file.close()

#fileds



def write_independently():
    #Before this, copy files from nodes to master using ./bmetriccollector.sh and run this code on master

    #settings
    metrics_file="metrics.txt"
    excel_file_path = "/home/" + getpass.getuser() + "/logs/metrics.xlsx" # this file should already be there with a sheet named nodes
    #assume csv file is also with the same nam in the same directory

    log_folder="/home/ubuntu/logs/" #on master
    test_name_format="alg-ALGORITHM_oth1-SCHEDCONF_scale-hpa_repamax-1_cpureq-1800_cpulim-3600_oth2-na_workloadtyp-sync_shp-off_rng-WORKLOAD_RATE_oth3-hold1_cpugov-ondemand_schint-5_battsize-1250_solarsize-1_oth4-OTHERSCONF_round-ROUND"
    NODES=["master", "homo1", "homo2", "homo3", "homo4", "homo5", "homo6", "homo7", "homo8", "homo9", "homo10"]
    

    ALGORITHMS=[ "greedy"] # idle local default random shortfaas greedy
    SCHEDCONFS=["z1006030-stick0.2"] #na plug1000000 z1006030-stick0.2
    OTHERSCONFS=["boot5-rep1-hicup2block30"]  # na boot5-rep1-hicup3block20 boot5-rep1-hicup2block30
    ROUNDS=["1"]
    WORKLOAD_RATES=["13"] # 33 13 14 31

    #RUN
    #read json data

    #per algorithm
    for algorithm in ALGORITHMS:
        #per round
        for round in ROUNDS:
            for schedconf in SCHEDCONFS:
                #add more filters here (before node)

                #per othersconf
                for othersconf in OTHERSCONFS:
                    #...
                    #per workload rate
                    for workload_rate in WORKLOAD_RATES:
                        #per node
                        for node in NODES:

                            #prepare test_name
                            
                            #replace algorithm
                            test_name= test_name_format.replace('ALGORITHM', algorithm)          
                            #replace round
                            test_name = test_name.replace('ROUND', round)
                            #replace schedconf
                            test_name = test_name.replace('SCHEDCONF', schedconf)
                            #replace othersconf
                            test_name = test_name.replace('OTHERSCONF', othersconf)
                            #replace workload rate
                            test_name = test_name.replace('WORKLOAD_RATE', workload_rate)
                            #add more formatting here
                            #...

                            
                            #metrics full path
                            if node == "master":
                                #master (master logs are in logs directory, not anotehr subdirectory)
                                metrics_full_path = log_folder  + '/' + node + '_' + test_name + '/' + metrics_file
                                master_done = True
                            else:
                                #nodes (nodes logs are in a subfolder under their name)
                                metrics_full_path = log_folder + node +  '/' + node + '_' + test_name + '/' + metrics_file
                                
                            try:
                                #read metrics.txt as json
                                with open(metrics_full_path, 'r') as f:
                                    my_dict = json.load(f)


                                print(f'node={node} metrx.txt read done')

                                #write

                                #call write (it calls scheduler and csv writer also)
                                write(my_dict, excel_file_path)

                                


                            except Exception as e:
                                print(f'******************read/write metrics failed={str(e)}')

                        #call write avg: average of all nodes, no matter if a node had metrics or not
                        print(f'write everage ')
                        #exclude master
                        worker_nodes = len(NODES) - 1
                        write_avg(worker_nodes, excel_file_path)
                        
                        print('write average done')

                        #NODES
                    #WORKLOAD_RATES
                #OTHERSCONFS
            #SCHEDCONFS
        #ROUND
    #ALGORITHMS
    return None

# write_independently()