import csv
import threading
import fire

#['test', 'time', 'dur', 'start', 'finish', 'name', 'power', 'down_min', 'down_%', 'cpu_avg', 'cpu_max', 'cpuUp_avg', 'cpuUp-max', 'cpu_freq_avg', 'mem_avg', 'mem_max', 'bw_sent(mb)', 'bw_recv(mb)', 'aplitter', 'app', 'created', 'sent_sum', 'sent%', 'code200_sum', 'code200%', 'code500', 'code502', 'code503', 'code-1', 'others', 'drop_sum', 'drop%', 'drop_b_sum', 'drop_b%', 'adm_avg', 'adm_max', 'qu_avg', 'qu_max', 'exec_avg', 'exec_max', 'rt_suc_avg', 'rt_suc__max', 'rt_suc_fail_avg', 'rt_suc_fail_max', 'useless', 'throu2', 'p0_suc', 'p25', 'p50', 'p75', 'p90', 'p95', 'p99', 'p99.9', 'p100', 'p0_suc_fail', 'p25', 'p50', 'p75', 'p90', 'p95', 'p99', 'p99.9', 'p100', 'detect_sum', 'detect_avg', 'detect_accuracy', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']

#read fields from a row in a csv file as dict
def read_fields_from_csv(csv_file_path, field_names, row_num=None, filters={}):
    # field_names = ['field1' , 'field2']
    #filters = {'fieldname1': 'filtervalue1', 'fieldname2': 'filtervalue2'}
    import csv
    lock = threading.Lock()

    picked_dict = {}
    breaker =False

    with lock:
        try:
            #read csv_file_path
            with open(csv_file_path, newline='') as csvfile:
                reader = csv.DictReader(csvfile)

                row_count = 1
                # iterate over the rows as dict
                for row in reader:

                    if breaker :
                        break
                    
                    row_count += 1
                    # Check row_count condition
                    if row_num is not None:
                        # reached the row_num
                        if row_count == row_num:
                    
                            # pick requested fields by their name
                            picked_dict = {}
                            for field_name in field_names:
                    
                                # Check if the field exists
                                if row.get(field_name):
                                    # add it to a dict
                                    picked_dict[field_name] = row[field_name]
                                else:
                                    print('field_names[' + field_name + '] does not exist in' + csv_file_path)
                                    picked_dict[field_name] = 'na'
                            
                            #stop searching
                            breaker=True

                    # Check row by filters
                    elif len(filters) > 0:

                        filter_passed = 0

                        # Check all filters in the row
                        for field, filter in filters.items():
                            
                            # field vs filter
                            if row[field] == filter:
                                filter_passed += 1

                            # if matched filters = request filters
                            if filter_passed == len(filters):

                                # pick field from this row. It only considers the first matching row.???
                                for field_name in field_names:
                                    # Check if the field exists
                                    if row.get(field_name):
                                        # add it to a dict
                                        picked_dict[field_name] = row[field_name]
                                    else:
                                        picked_dict[field_name] = 'na'
                                        print('field_names[' + field_name + '] does not exist/have a value in ' + csv_file_path)
                                
                                # stop searching
                                breaker =True
                                break
                    else:
                        print('either row_num or filters is required for searching.')


                return picked_dict
        except Exception as e:
            print('Reading file failed.\n' + str(e))
            return str(e)


import threading
import datetime
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE


def write_a_row_to_excel(excel_file_path, sheet_name, data, row_num=None):
    #data = {{'A': 1000}, {'D': 'ali'}}
    lock = threading.Lock()
    with lock:
        #load sheet
        try:
            wb = load_workbook(filename = excel_file_path)
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                err = 'sheetname= ' + sheet_name + ' does not exist.'
                print(err)
                return err
        except Exception as e:
            return str(e)

        #select the last row number if no row_num is given
        if row_num == None:
            #last written row   index
            max_row = sheet.max_row
            #current row index
            row_num = max_row + 1

        #add data
        for column_name, value in data.items():
            if '_' in str(value):
                sheet[column_name + str(row_num)] = str(value)
            elif '.' in value:
                sheet[column_name + str(row_num)] = float(value)
            else:
                sheet[column_name + str(row_num)] = int(value)
        #save
        wb.save(filename = excel_file_path)
        wb.close()


by_test_name = False

index1_req_st = 9
index1_req_end = 9
index2_th_flask = 4
index2_th_model = 4
index3_rowstart = 108
index4_rowjump = 0
suffix_name = 'w2_' +'sync-pi4-4gb-rev1.1-cpu-rep1-req'
prefix_name = '_2'
#sheet name?
sheet_name = 'all-together-sync'

for i in range(index1_req_st, index1_req_end + 1):
    #[input]
    
    #read from CSV
    csv_file_path = '/home/ubuntu/logs/metrics.csv'
    # csv_row_num = 2
    filters={'test': suffix_name + str(i) + '-th' + str(index2_th_flask) + '_' + str(index2_th_model) + prefix_name}
    # filter = {'time': '2022-09-25 06:20:15.937861+00:00'}
    #write to excel 
    excel_file_path = '/home/ubuntu/profiling.xlsx'
    
    #row to write?

    excel_row_num = index3_rowstart + ((i - index1_req_st) * index4_rowjump)



    #fields in csv file
    read_list = ['test', 'power', 'cpu_avg', 'cpu_max', 'cpu_freq_avg', 'mem_avg','bw_sent(mb)', 'bw_recv(mb)', 'created', 'code200%',	'exec_avg', 'exec_max',
                'rt_suc_avg', 'rt_suc_max', 'throu2', 'detect_avg', 'detect_accuracy', 'dur',
                'p0_suc', 'p25_suc', 'p50_suc', 'p75_suc', 'p90_suc', 'p95_suc', 'p99_suc', 'p99.9_suc', 'p100_suc'
    ]

    #[read]
    # res = read_fields_from_csv(csv_file_path, read_list,row_num= csv_row_num)
    res = read_fields_from_csv(csv_file_path, read_list, None, filters)

    #[write]
    #columns in xlsx file metrics.xlsx and their expected type of value
    xlsx_map = {'AC': 'test', 'I': 'power', 'J': 'cpu_avg', 'K': 'cpu_max', 'L': 'cpu_freq_avg', 'M': 'mem_avg', 'N': 'bw_sent(mb)', 'O': 'bw_recv(mb)', 
                'P': 'created', 'Q': 'code200%', 'R': 'exec_avg', 'S': 'exec_max', 'T': 'rt_suc_avg', 'U': 'rt_suc_max', 'V': 'throu2', 
                'W': 'detect_avg', 'X': 'detect_accuracy', 'AB': 'dur',
                'AD': 'p0_suc', 'AE': 'p25_suc', 'AF': 'p50_suc', 'AG': 'p75_suc', 'AH': 'p90_suc', 'AI': 'p95_suc', 
                'AJ': 'p99_suc', 'AK': 'p99.9_suc', 'AL': 'p100_suc'}

    #create a dict by read data
    data ={}
    for excel_col, corrs_csv_field in xlsx_map.items(): 
        if res.get(corrs_csv_field):
            data[excel_col] = res[corrs_csv_field]
        else:
            print('key=' + corrs_csv_field + ' not found')


    wrt = write_a_row_to_excel(excel_file_path, sheet_name, data, excel_row_num)

# if __name__ == '__main__':
#   fire.Fire()